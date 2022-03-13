import sys,os
import openpyxl
from Connect_to_Boson import debugging,Get_PxsshObj,get_Device
from backendservercheck import getUsers
#from tasks import findFFT, FNDStatus
import logging
import time
import string
import re


start_time = time.time()
debug = True        # Enable debugint messages
debugToFile = False  #
'''Boson Parameters'''
bosonServer   = 'devices.dev.siq.sleepnumber.com'
bosonUser     = 'hassan'
bosonPassword = 'nura6719'
bosonPort     = '6022'
bosonPrompt   = '[#|$]'
bosonTimeout  =  10
bosonSSHKEY   ='/home/hassan/sshkey/mypkey4'
'''Backendserver Parameters'''
backendServer_URL = "https://prod-be-api.sleepiq.sleepnumber.com/bam/rest/sn/v1/"
backendServer_UserName = "hassan@bamlabs.com"
backendServer_Password = "Test1234"



debugging(debug,debugToFile) # Call debug method

child = Get_PxsshObj(bosonServer,bosonUser,bosonPassword,bosonPort,bosonPrompt, bosonTimeout, bosonSSHKEY) # Get pexpect onject after connecting to Boson server
if child == 1:
    print("Unable to establish connection with Boson server exiting program:") # failed to connect to Boson no need to
    sys.exit()                                                                 # continue
#----------------------------------------------------------------------------------------------------------------------
#Input file
getMAC = openpyxl.load_workbook('MAC-list3.xlsx')
inputSheet = getMAC.get_sheet_by_name('MAC-Sheet')
inputSheet = getMAC.active

# Output file
wb = openpyxl.Workbook()
outPutSheet = wb.get_sheet_names()
outPutSheet = wb.active

outPutSheet.title = "FND"
outPutSheet['A1'] = "MAC"
outPutSheet['B1'] = "Server"
outPutSheet['C1'] = "User"
outPutSheet['D1'] = "Notes"
outPutSheet['E1'] = "FND Vers"
outPutSheet['F1'] = "B"
outPutSheet['G1'] = "O"
outPutSheet['H1'] = "F"
outPutSheet['I1'] = "R"
outPutSheet['J1'] = "A"
outPutSheet['K1'] = "FW"




#outputfile= openpyxl.load_workbook('output.xlsx')


for row in range(2,inputSheet.max_row + 1): # Last value excluded so add 1
    MAC= inputSheet['A' + str(row)].value
    if MAC == None:
        print("Empty row skipping: ")
        logging.debug("Empty row skipping:")
        outPutSheet['A' + str(row)].value = "Empty row"
        outPutSheet['B' + str(row)].value = "---"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue
    if not ((MAC.startswith("CC04B4") or MAC.startswith("cc04b4") or MAC.startswith("64DBA0") or MAC.startswith("64dba0")) and (len(MAC)==12)):  # | "cc04b4" | "640000"):
        print("Incorrect format for MAC address skipping: ")
        print(len(MAC))
        outPutSheet['A' + str(row)].value = "Incorrect format for MAC   address"
        outPutSheet['B' + str(row)].value = "---"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue

    MAC = MAC.upper()  # MAC addresses stored with upper case in backend server
    device = get_Device(child, MAC, backendServer_URL, backendServer_UserName, backendServer_Password)

    if device == 1:
        print("Device does not exists in the backend server: Skipping.")
        outPutSheet['A' + str(row)].value = "Device does not exists in the backend server"
        outPutSheet['B' + str(row)].value = "Skip"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue
    if device == 2:
        print("Device is offline: Skipping.")
        outPutSheet['A' + str(row)].value = "Device is offline"
        outPutSheet['B' + str(row)].value = "Skip"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue
    if device == 3:
        print("Device taking too long to recover: skipping")
        outPutSheet['A' + str(row)].value = "Device taking too long to recover"
        outPutSheet['B' + str(row)].value = "Skip"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue
    if device == 4:
        print("Error end of file expection : skipping.")
        outPutSheet['A' + str(row)].value = "Err or end of file expection"
        outPutSheet['B' + str(row)].value = "Skip"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue
    if device == 5:
        print("Error TIMEOUT : Skipping.")
        outPutSheet['A' + str(row)].value = "Error TIMEOUT"
        outPutSheet['B' + str(row)].value = "Skip"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue
    if device == 6:
        print("Unkwon execpetion : Skipping.")
        outPutSheet['A' + str(row)].value = "Unkwon execpetion"
        outPutSheet['B' + str(row)].value = "Skip"
        outPutSheet['C' + str(row)].value = "---"
        outPutSheet['D' + str(row)].value = "---"
        outPutSheet['E' + str(row)].value = "---"
        outPutSheet['F' + str(row)].value = "---"
        outPutSheet['G' + str(row)].value = "---"
        outPutSheet['H' + str(row)].value = "---"
        outPutSheet['I' + str(row)].value = "---"
        outPutSheet['J' + str(row)].value = "---"
        outPutSheet['K' + str(row)].value = "---"
        continue
    if device == 0:   # here where the work begin for each connected device
        print("Connected to device :" + str(MAC))
        #Add the MAC address to the output file
        outPutSheet['A' + str(row)].value = str(MAC)
        
        #Get the server name
        child.sendline("cat /bam/etc/bam.conf ")
        child.expect(":~#")
        serverNmae = child.before
        serverNmae = serverNmae.decode()
        serverNmae = re.search(r'(^blohURL\s=\shttps://)(.*?)(\.)',serverNmae, re.M)
        outPutSheet['B' + str(row)].value = serverNmae.group(2)

        #Get user information
        outPutSheet['C' + str(row)].value = getUsers(MAC, backendServer_URL, backendServer_UserName, backendServer_Password)

        #Not to be implemented
        #outPutSheet['D' + str(row)].value =XXXXXX

        #Get the foundation version
        child.sendline("bio plgg; bio mfvr; baml")
        child.expect(":~#")
        logging.debug(child.after)
        logging.debug(child.before)
        logging.debug(child.buffer)

        #/////////////////////////////////////////

        fndVersion = child.before
        fndVersion = fndVersion.decode()
        print(fndVersion)

        findFNDversion = re.search(r'(^FND version<)(0x[0-9A-Za-z]{8})(>\r$)', fndVersion, re.M)
        if findFNDversion != None:
            version = findFNDversion.group(2)
            print(findFNDversion.group(2))
            # to do writ to to file
        else:
            version = "N"
            print("error no version found")
        outPutSheet['E' + str(row)].value =  str(version)


        #/////////////////////////////////////////////

        findFNDversion = re.search(r'(^\s\slog\slevel<)(\d)(>)', fndVersion, re.M)
        if findFNDversion != None:
            logLevel = findFNDversion.group(2)
            if logLevel == "1":
                outPutSheet['F' + str(row)].value = "Y"
            if logLevel == "0":
                child.sendline("baml 1 1 month")
                child.expect(":~#")
                outPutSheet['F' + str(row)].value = "N->Y"



        child.sendline("logout")
        child.expect("closed.")




        #findFFT(MAC, row,outPutSheet,child,intiate)
        #print("intiate before::::" + intiate[0])
        #FNDStatus(MAC, row,outPutSheet,child,intiate)
        #print("intiate before::::" + intiate[0])
        print(MAC)
        print("--- %s seconds ---" % (time.time() - start_time))









print("--- %s seconds ---" % (time.time() - start_time))
wb.save('empty_book.xlsx')
print("--- %s seconds ---" % (time.time() - start_time))


#Call
