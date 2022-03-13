import sys,os
import openpyxl
from Connect_to_Boson import debugging,Get_PxsshObj,get_Device
from tasks import findFFT, FNDStatus
import logging
import time
import string



start_time = time.time()
debug = True        # Enable debugint messages
debugToFile = False  #
'''Boson Parameters'''
bosonServer   = 'devices.dev.siq.sleepnumber.com'
bosonUser     = 'hassan'
bosonPassword = 'nura6719'
bosonPort     = '6022'
bosonPrompt   = '[#|$]'
bosonTimeout  =  10
bosonSSHKEY   ='/home/osama/sshkey/mypkey4'
'''Backendserver Parameters'''
backendServer_URL = "https://prod-be-api.sleepiq.sleepnumber.com/bam/rest/sn/v1/"
backendServer_UserName = "hassan@bamlabs.com"
backendServer_Password = "Test1234"



debugging(debug,debugToFile) # Call debug method

child = Get_PxsshObj(bosonServer,bosonUser,bosonPassword,bosonPort,bosonPrompt, bosonTimeout, bosonSSHKEY) # Get pexpect onject after connecting to Boson server
if child == 1:
    print("Unable to establish connection with Boson server exiting program:") # failed to connect to Boson no need to
    sys.exit()                                                                 # continue
#----------------------------------------------------------------------------------------------------------------------
#Input file
getMAC = openpyxl.load_workbook('MAC-list3.xlsx')
inputSheet = getMAC.get_sheet_by_name('MAC-Sheet')
inputSheet = getMAC.active

# Output file
wb = openpyxl.Workbook()
outPutSheet = wb.get_sheet_names()
outPutSheet = wb.active
'''
outPutSheet.title = "FFT"
outPutSheet['A1'] = "MAC"
outPutSheet['B1'] = "Status"
'''
intiate=["true"]




#outputfile= openpyxl.load_workbook('output.xlsx')


for row in range(2,inputSheet.max_row + 1): # Last value excluded so add 1
    MAC= inputSheet['A' + str(row)].value
    if MAC == None:
        print("Empty row skipping: ")
        logging.debug("Empty row skipping:")
        outPutSheet['A' + str(row)].value = "Empty row"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue
    if not ((MAC.startswith("CC04B4") or MAC.startswith("cc04b4") or MAC.startswith("64DBA0") or MAC.startswith("64dba0")) and (len(MAC)==12)):  # | "cc04b4" | "640000"):
        print("Incorrect format for MAC address skipping: ")
        print(len(MAC))
        outPutSheet['A' + str(row)].value = "Incorrect format for MAC   address"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue

    MAC = MAC.upper()  # MAC addresses stored with upper case in backend server
    device = get_Device(child, MAC, backendServer_URL, backendServer_UserName, backendServer_Password)

    if device == 1:
        print("Device does not exists in the backend server: Skipping.")
        outPutSheet['A' + str(row)].value = "Device does not exists in the backend server"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue
    if device == 2:
        print("Device is offline: Skipping.")
        outPutSheet['A' + str(row)].value = "Device is offline"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue
    if device == 3:
        print("Device taking too long to recover: skipping")
        outPutSheet['A' + str(row)].value = "Device taking too long to recover"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue
    if device == 4:
        print("Error end of file expection : skipping.")
        outPutSheet['A' + str(row)].value = "Error end of file expection"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue
    if device == 5:
        print("Error TIMEOUT : Skipping.")
        outPutSheet['A' + str(row)].value = "Error TIMEOUT"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue
    if device == 6:
        print("Unkwon execpetion : Skipping.")
        outPutSheet['A' + str(row)].value = "Unkwon execpetion"
        outPutSheet['B' + str(row)].value = "Skipping"
        continue
    if device == 0:
        print("Connected to device :" + str(MAC))
        findFFT(MAC, row,outPutSheet,child,intiate)
        #print("intiate before::::" + intiate[0])
        #FNDStatus(MAC, row,outPutSheet,child,intiate)
        #print("intiate before::::" + intiate[0])
        print(MAC)
        print("--- %s seconds ---" % (time.time() - start_time))










wb.save('empty_book.xlsx')
print("--- %s seconds ---" % (time.time() - start_time))


#Call
