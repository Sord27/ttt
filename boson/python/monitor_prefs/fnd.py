#------- Script Legend:--------------------------------
'''
ER     = Empty row
IF     = Incorrect format for MAC   address
DNSIBS = Device does not exists in the backend server
DOL    = Device is offline
DTTLTR = Device taking too long to recover
EEOFE  = Error end of file expection
ET     = Error TIMEOUT
UE     = Unkwon execpetion"
N->Y   = Log level 0 and enabled to 1
#------ Foundation Legend:--------------------------------
FFNA   = Firmware File not accessable
DONL    = device on line have access to Boson
FONL    =  Foundation on lin
FOFL    =  Foundation off lin
-       = Don't care
Auto    = Automatic reset
'''







import sys,os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from Connect_to_Boson import debugging,Get_PxsshObj,get_Device
from backendservercheck import getUsers, findBeServere
#from tasks import findFFT, FNDStatus
import logging
import time
import string
import re
from datetime import date
import calendar

font = Font(name='Calibri',size=11, bold=True,italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
border = Border( left=Side(border_style=None, color='FF000000'),
                 right=Side(border_style=None,color='FF000000'),
                 top=Side(border_style=None, color='FF000000'),
                 bottom=Side(border_style=None,color='FF000000'),
                 diagonal=Side(border_style=None,color='FF000000'),
                 diagonal_direction=0, outline=Side(border_style=None, color='FF000000'),
                 vertical=Side(border_style=None, color='FF000000'),
                 horizontal=Side(border_style=None, color='FF000000'))
alignment=Alignment(horizontal='general',
                   vertical='bottom',
                    text_rotation=0,
                   wrap_text=False,
                   shrink_to_fit=False,
                    indent=0)
number_format = 'General'
protection = Protection(locked=True, hidden=False)



start_time = time.time()
debug = False #True        # Enable debugint messages
debugToFile = False  #
'''Boson Parameters'''
bosonServer   = 'devices.dev.siq.sleepnumber.com'
bosonUser     = 'hassan'
bosonPassword = 'nura6719'
bosonPort     = '6022'
bosonPrompt   = '[#|$]'
bosonTimeout  =  10
bosonSSHKEY   ='/home/osama/sshkey/mypkey4'
'''Backendserver Parameters'''
#backendServer_URL = "https://prod-be-api.sleepiq.sleepnumber.com/bam/rest/sn/v1/"
#backendServer_URL = "https://circle2-be-api.circle.siq.sleepnumber.com/bam/rest/sn/v1/"
#backendServer_UserName = "hassan@bamlabs.com"
#backendServer_Password = "Test1234"



debugging(debug,debugToFile) # Call debug method

child = Get_PxsshObj(bosonServer,bosonUser,bosonPassword,bosonPort,bosonPrompt, bosonTimeout, bosonSSHKEY) # Get pexpect onject after connecting to Boson server
if child == 1:
    print("Unable to establish connection with Boson server exiting program:") # failed to connect to Boson no need to
    sys.exit()                                                                 # continue
#----------------------------------------------------------------------------------------------------------------------
#Input file
devicenumber = 1
getMAC = openpyxl.load_workbook('store/'+'MAC-list3.xlsx') #??????????location
inputSheet = getMAC.get_sheet_by_name('MAC-Sheet')
inputSheet = getMAC.active

# Output file
wb = openpyxl.Workbook()
outPutSheet = wb.get_sheet_names()
outPutSheet = wb.active

outPutSheet.title = "FND"

#Server
outPutSheet['A1'] = "MAC"
outPutSheet['B1'] = "Server/status"
outPutSheet['C1'] = "First-N"
outPutSheet['D1'] = "Last-N"
outPutSheet['E1'] = "Login"
outPutSheet['F1'] = "Notes"
outPutSheet['G1'] = "FND Vers"
outPutSheet['H1'] = "B"
#Foundation
outPutSheet['I1'] = "O"
outPutSheet['J1'] = "F"
outPutSheet['K1'] = "R"
outPutSheet['L1'] = "A"
outPutSheet['M1'] = "FW-BIN"
outPutSheet['N1'] = "Match"
#Responsive Air
outPutSheet['O1'] = "RAR-EN"
outPutSheet['P1'] = "RAL-EN"
outPutSheet['Q1'] = "RAR-ACT"
outPutSheet['R1'] = "RAL-ACT"
#UnderBed Lighting
outPutSheet['S1'] = "UBL-EN"
outPutSheet['T1'] = "UBL-ON"
outPutSheet['U1'] = "UBL-OFF"




for row in range(2,inputSheet.max_row + 1): # Last value excluded so add 1
    MAC= inputSheet['A' + str(row)].value
    # Make sure the input value is not an empty string
    if MAC == None:
        print("Empty row skipping device #: " + str(devicenumber))
        logging.debug("Empty row skipping device #: " + str(devicenumber))
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "ER"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"
        devicenumber += 1
        continue

    # Make sure the MAC is in the correct format and start with correct initial  6 digits for Select Comfort
    if not ((MAC.startswith("CC04B4") or MAC.startswith("cc04b4") or MAC.startswith("64DBA0") or MAC.startswith("64dba0")) and (len(MAC)==12)):  # | "cc04b4" | "640000"):
        print("Incorrect format for MAC address skipping device #: " + str(devicenumber))
        #print(len(MAC)) # remove
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "IF"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"
        devicenumber += 1
        continue

    MAC = MAC.upper()  # MAC addresses stored with upper case in backend server

    # Find the device server and set the URL, user and passsowrd
    credential = findBeServere(MAC)
    if len(credential) == 3:
        backendServer_URL = credential[0]
        backendServer_UserName = credential[1]
        backendServer_Password = credential[2]
    else: # the device could not be located in any known server, process will be skipped.
        devicenumber += 1  #
        child.sendline("logout")
        child.expect("closed.")
        continue


    # Attempting to access the deice through SSH through Boson.
    device = get_Device(child, MAC, backendServer_URL, backendServer_UserName, backendServer_Password)
    if device == 1:
        print("Device does not exists in the backend server Skipping device #: " + str(devicenumber))
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "DNSIBS"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"
        devicenumber += 1
        continue
    if device == 2:
        print("Device is offline: Skipping device #: " + str(devicenumber))
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "DOL"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"
        devicenumber += 1
        continue
    if device == 3:
        print("Device taking too long to recover skipping device #: " + str(devicenumber))
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "DTTLTR"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"

        devicenumber += 1
        continue
    if device == 4:
        print("Error end of file expection skipping device #: " + str(devicenumber))
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "EEOFE"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"
        devicenumber += 1
        continue
    if device == 5:
        print("Error TIMEOUT Skipping device #: " + str(devicenumber))
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "ET"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"
        devicenumber += 1
        continue
    if device == 6:
        print("Unkwon execpetion Skipping device #: " + str(devicenumber))
        outPutSheet['A' + str(row)].value = str(MAC)
        outPutSheet['B' + str(row)].value = "UE"
        outPutSheet['C' + str(row)].value = "-"
        outPutSheet['D' + str(row)].value = "-"
        outPutSheet['E' + str(row)].value = "-"
        outPutSheet['F' + str(row)].value = "-"
        outPutSheet['G' + str(row)].value = "-"
        outPutSheet['H' + str(row)].value = "-"
        outPutSheet['I' + str(row)].value = "N"
        outPutSheet['J' + str(row)].value = "-"
        outPutSheet['K' + str(row)].value = "-"
        outPutSheet['L' + str(row)].value = "-"
        outPutSheet['M' + str(row)].value = "-"
        outPutSheet['N' + str(row)].value = "-"
        outPutSheet['O' + str(row)].value = "-"
        outPutSheet['P' + str(row)].value = "-"
        outPutSheet['Q' + str(row)].value = "-"
        outPutSheet['R' + str(row)].value = "-"
        outPutSheet['S' + str(row)].value = "-"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"
        devicenumber += 1
        continue

    # Connection to the device established.
    if device == 0:   # here where the work begin for each connected device
        print("Connected to device # " + str(devicenumber) +": " + str(MAC))
        outPutSheet['A' + str(row)].value = str(MAC)  # Add the MAC to the output file
        outPutSheet['I' + str(row)].value = "Y"       # The status = connected
        
        #Get the server name -------------------------------------------------------
        child.sendline("cat /bam/etc/bam.conf ")
        child.expect(":~#")
        serverNmae = child.before
        serverNmae = serverNmae.decode()
        serverNmae = re.search(r'(^blohURL\s=\shttps://)(.*?)(\.)',serverNmae, re.M)
        outPutSheet['B' + str(row)].value = serverNmae.group(2)  # Add the server to the output file

        # Get user information -----------------------------------------------------
        userInfo = []
        userInfo = getUsers(MAC, backendServer_URL, backendServer_UserName, backendServer_Password)
        if len(userInfo) == 3:
            outPutSheet['C' + str(row)].value = userInfo[0] # User info first name
            outPutSheet['D' + str(row)].value = userInfo[1] # Last name
            outPutSheet['E' + str(row)].value = userInfo[2] # login

        #Note to be implemented ----------------------------------------------------
        outPutSheet['F' + str(row)].value ="-"

        #Get the running foundation version ----------------------------------------
        child.sendline("bio plgg; bio mfvr; baml")
        child.expect(":~#")
        logging.debug(child.after)
        logging.debug(child.before)
        logging.debug(child.buffer)
        fndVersion = child.before
        fndVersion = fndVersion.decode()
        findFNDversion = re.search(r'(^FND version<)(0x[0-9A-Za-z]{8})(>\r$)', fndVersion, re.M) # look for the FND ver
        if findFNDversion != None:
            runningFndVersion = findFNDversion.group(2)
            outPutSheet['G' + str(row)].value = str(runningFndVersion)
            outPutSheet['J' + str(row)].value = "Y"  # FND On Line
            outPutSheet['K' + str(row)].value = "-"     # Don't care
            outPutSheet['L' + str(row)].value = "Y"  # Auto reset, Yes
        else:
            runningFndVersion = "N"
            outPutSheet['G' + str(row)].value = "N"
            outPutSheet['J' + str(row)].value = "N"  # FND Off line
            outPutSheet['K' + str(row)].value = "-"     # Don't care
            outPutSheet['L' + str(row)].value = "-"     # Don't care
            #outPutSheet['M' + str(row)].value = "-"     # Don't care
            #outPutSheet['N' + str(row)].value = "-"     # Don't care
           # devicenumber += 1#??????????????????????????????????????????????????/
            #child.sendline("logout")#???????????????????????????????????????????????
            #child.expect("closed.")#???????????????????????????????????????????????????
            #continue #????????????????????????????????????????????????/  # IF the FND is dead jump out of the loop

        # Check the log level and set it if it is 0 --------------------------------------------------
        findFNDversion = re.search(r'(^\s\slog\slevel<)(\d)(>)', fndVersion, re.M)
        if findFNDversion != None:
            logLevel = findFNDversion.group(2)
            if logLevel == "1":
                outPutSheet['H' + str(row)].value = "Y"
            if logLevel == "0":
                child.sendline("baml 1 1 month")
                child.expect(":~#")
                outPutSheet['H' + str(row)].value = "N->Y" # found as 0 but enabled to 1


        # Check the firmware version in the bin file -------------------------------------
        child.sendline("cd /bam/firmware; ls")
        child.expect("firmware#")
        fndFirmwareVersion = child.before
        fndFirmwareVersion = fndFirmwareVersion.decode()
        fndFirmwareVersionobj = re.search(r'(foundation-device-)([0-9A-Za-z]{3}35[0-9A-Za-z]{3})(\.bin\r\n)', fndFirmwareVersion, re.M)
        if fndFirmwareVersionobj != None:
            ####print("\n2")
            ####print(fndFirmwareVersionobj.group(2))
            #####print("\n3")
            storedFndVersion = "0x" + fndFirmwareVersionobj.group(2)
            #storedFndVersion = fndFirmwareVersionobj
            #####print(storedFndVersion)
            outPutSheet['M' + str(row)].value = str(storedFndVersion)
            # to do writ to to file
        else:
            storedFndVersion = "N"
            outPutSheet['M' + str(row)].value = "FFNA"
        if str(storedFndVersion) == str(runningFndVersion):
            outPutSheet['N' + str(row)].value = "Y"
        else:
            outPutSheet['N' + str(row)].value = "N"

        # --------------------------Responsive Air------------------------------------------------------

        getDate = time.strftime("%m/%d/%y")  # check
        my_date = date.today()
        dayOftheweek = calendar.day_name[my_date.weekday()]
        dayOftheweek = dayOftheweek.lower()
        child.sendline("cat /bam/etc/preferences")
        child.expect('/bam/firmware#')
        preferences = child.before
        preferences = preferences.decode()
        logging.debug(preferences)
        # Check if RA right side is on
        ra_Right = re.search(r'(^ra_right_enable=)(.*)(\r\r$)', preferences, re.M)
        ra_Right_Status = ra_Right.group(2)
        if ra_Right_Status == "true":
            outPutSheet['O' + str(row)].value = "Y" # Writ to file
        else:
            outPutSheet['O' + str(row)].value = "N"
        # Check if RA left side is on
        ra_Left = re.search(r'(^ra_left_enable=)(.*)(\r\r$)', preferences, re.M)
        ra_Left_Status = ra_Left.group(2)
        if ra_Left_Status == "true":
            outPutSheet['P' + str(row)].value = "Y"
        else:
            outPutSheet['P' + str(row)].value = "N"
        # Check if there is an activity in either side of the RA.
        if ra_Right_Status == "true" or ra_Left_Status == "true":
            child.sendline("bio rad")
            child.expect('/bam/firmware#')
            ra_activities = child.before
            ra_activities = ra_activities.decode()
            if ra_Right_Status == "true":
                # Find all right-side RA activities for today
                find_ra_Right_activities = re.compile(re.escape(getDate) + '(\s.*\sR\sadj\s+\d+\-\>\s\d+)')
                ra_Right_activities = find_ra_Right_activities.findall(ra_activities)
                ra_Right_activities_count = len(ra_Right_activities)
                outPutSheet['Q' + str(row)].value = str(ra_Right_activities_count) # Writ to file

            else:
                outPutSheet['Q' + str(row)].value = "-"
                # Find all left-side RA activities for today
            if ra_Left_Status == "true":
                find_ra_Left_activities = re.compile(re.escape(getDate) + '(\s.*\sL\sadj\s+\d+\-\>\s\d+)')
                ra_Left_activities = find_ra_Left_activities.findall(ra_activities)
                ra_Left_activities_count = len(ra_Left_activities)
                outPutSheet['R' + str(row)].value = str(ra_Left_activities_count)
            else:
                outPutSheet['R' + str(row)].value = "-"

     # --------------------------Underbed Lighting ------------------------------------------------------

    underBedLight = re.search(r'(^ul_auto_enabled=)(.*)(\r\r\n)', preferences, re.M)
    if underBedLight.group(2) == "true":
        outPutSheet['S' + str(row)].value = "Y"  #print("\nnumber of ons %d",NumberoflightTurnedOn)
        child.sendline("bio ul_logs")
        child.expect('/bam/firmware#')
        ul_logs = child.before
        ul_logs = ul_logs.decode()
        # logging.debug(preferences)
        lightTurnedOn = re.compile(re.escape(getDate) + '(\s\d\d:\d\d:\d\d\sset underbed to 1 for\s)')
        NumberoflightTurnedOn = lightTurnedOn.findall(ul_logs)
        NumberoflightTurnedOn = len(NumberoflightTurnedOn)
        outPutSheet['T' + str(row)].value = str(NumberoflightTurnedOn)
        lightTurnedOff = re.compile(re.escape(getDate) + '(\s\d\d:\d\d:\d\d\sturned off underbed)')
        NumberoflightTurnedOff = lightTurnedOff.findall(ul_logs)
        NumberoflightTurnedOff = len(NumberoflightTurnedOff)
        outPutSheet['U' + str(row)].value = str(NumberoflightTurnedOff)


    else:
        outPutSheet['S' + str(row)].value = "N"
        outPutSheet['T' + str(row)].value = "-"
        outPutSheet['U' + str(row)].value = "-"









    child.sendline("logout")
    child.expect("closed.")
    devicenumber += 1



    





'''
    col = outPutSheet.column_dimensions['B']
    col.alignment = Alignment(horizontal='center')
    a1 = outPutSheet['B1']
    a1.font = Font(bold=True)
    a1.alignment = Alignment(horizontal="center")
'''




#print("--- %s seconds ---" % (time.time() - start_time))

for num in range(0,18):
    index= string.ascii_uppercase[num]
    #index = index + str(1)
   # print("index = " + index )

    col = outPutSheet.column_dimensions[index]
    col.alignment = Alignment(horizontal='center')
    a1 = outPutSheet[index + str(1)]
    a1.font = Font(bold=True)
    a1.alignment = Alignment(horizontal="center")

#----------------------
for col in outPutSheet.columns:
    max_length = 0
    column = col[0].column  # Get the column name
    for cell in col:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1  #1.2
    outPutSheet.column_dimensions[column].width = adjusted_width


#-------------------------




timestr = time.strftime("%Y%m%d-%H%M%S")
fndoutfile= "fnd" + timestr + ".xlsx"
wb.save('store/output/'+ fndoutfile) #?????????????location
#wb.save('empty_book.xlsx')
print("Time to complete  =  %s seconds ---" % (time.time() - start_time))


#Call
